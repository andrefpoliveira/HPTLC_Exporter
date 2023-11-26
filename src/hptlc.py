import re, os, json

try:
    from src.reader import Reader
except:
    from reader import Reader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM

def rgb_to_hex(rgb):
    return '#%02x%02x%02x' % rgb

def rgba_to_hex(rgba):
    alpha = rgba[3]
    rgb = (int((1 - alpha) * 255 + alpha * x) for x in rgba[:3])
    return rgb_to_hex(rgb)

class HptlcReader(Reader):
    def __init__(self, filename, output="Teste.xlsx"):
        super().__init__(filename)
        self.output = output

    def extract_info(self) -> dict:
        tables = {}
        tracks = [x.strip() for x in re.findall(r"Track \d+, [^\n]+", self.text)]
        for id, track in enumerate(tracks):

            current_text = self.text[self.text.find(track):]
            current_text = current_text.split("Assigned substance")[1]

            if id < len(tracks) - 1:
                current_text = current_text.split(tracks[id + 1])[0]

            tables[track] = self.text_to_array(current_text.strip())

        return tables
    
    def text_to_array(self, text: str) -> list:
        table = []
        lines = text.split("\n")

        for line in lines:
            if not line.split(): continue
            peak = line.split()[0]

            m = re.findall(r"^\d", peak)
            if len(m) == 0: continue

            if "Peak deleted" in line: continue
            if len(line.split()) <= 10: continue
            
            peak, start_rf, start_h, max_rf, max_h, max_perc, end_rf, end_h, area, area_perc, *substance = line.split()
            table.append([peak, start_rf, start_h, max_rf, max_h, max_perc, end_rf, end_h, area, area_perc, " ".join(substance)])

        return table
    
    def build_sheet(self, sheet, groups, samples, substances):
        table_size = len(substances)
        for x in range(len(groups)):
            THIN_BORDER = Border(
                left=Side(border_style=BORDER_MEDIUM, color=groups[x]["main_color"]),
                right=Side(border_style=BORDER_MEDIUM, color=groups[x]["main_color"]),
                top=Side(border_style=BORDER_MEDIUM, color=groups[x]["main_color"]),
                bottom=Side(border_style=BORDER_MEDIUM, color=groups[x]["main_color"])
            )

            # Setup the group Header
            sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=2, end_row=2)
            sheet.cell(row=2, column=2 + x * (7 + 1)).value = groups[x]["name"].upper()
            sheet.cell(row=2, column=2 + x * (7 + 1)).font = sheet.cell(row=2, column=2 + x * (7 + 1)).font.copy(bold=True)
            sheet.cell(row=2, column=2 + x * (7 + 1)).alignment = sheet.cell(row=2, column=2 + x * (7 + 1)).alignment.copy(horizontal="center")
            sheet.cell(row=2, column=2 + x * (7 + 1)).fill = PatternFill(start_color=groups[x]["main_color"],end_color=groups[x]["main_color"], fill_type='solid')

            # Setup the table header
            for y in range(samples):

                for col in range(2 + x * (7 + 1), 2 + x * (7 + 1) + 7):
                    for row in range(4 + y * (table_size + 4), 7 + table_size + y * (table_size + 4)):
                        sheet.cell(row, col).border = THIN_BORDER

                sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=4 + y * (table_size + 4), end_row= 4 + y * (table_size + 4))
                sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).value = f"{groups[x]['acronym']}{x * samples + y + 1}"
                sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).font = sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).font.copy(bold=True)
                sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).alignment = sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).alignment.copy(horizontal="center")
                sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).fill = PatternFill(start_color=groups[x]["secondary_color"],end_color=groups[x]["secondary_color"], fill_type='solid')
                sheet.cell(row=4 + y * (table_size + 4), column=2 + x * (7 + 1)).border = THIN_BORDER

                headers = ["Lípidos", "Área", "%", "Área", "%", "Média", "Desvio Padrão"]
                for i in range(7):
                    sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).value = headers[i]
                    sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).font = sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).font.copy(bold=True)
                    if i != 0:
                        sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).alignment = sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).alignment.copy(horizontal="center")
                    sheet.cell(5 + y * (table_size + 4), column=2 + x * (7 + 1) + i).fill = PatternFill(start_color=groups[x]["tertiary_color"],end_color=groups[x]["tertiary_color"], fill_type='solid')

                for id, s in enumerate(substances):
                    row = 6 + id + y * (table_size + 4)
                    column = 2 + x * (7 + 1)
                    sheet.cell(row, column).value = s

                    sheet.cell(row, column + 2).value = f"=(${get_column_letter(column+1)}{row}/{get_column_letter(column+1)}${row + table_size - id})*100"
                    sheet.cell(row, column + 2).alignment = sheet.cell(row, column + 2).alignment.copy(horizontal="center")

                    sheet.cell(row, column + 4).value = f"=(${get_column_letter(column+3)}{row}/${get_column_letter(column+3)}{row + table_size - id})*100"
                    sheet.cell(row, column + 4).alignment = sheet.cell(row, column + 4).alignment.copy(horizontal="center")

                    sheet.cell(row, column + 5).value = f"=AVERAGE(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})"
                    sheet.cell(row, column + 5).alignment = sheet.cell(row, column + 5).alignment.copy(horizontal="center")

                    sheet.cell(row, column + 6).value = f"=_xlfn.STDEV.P(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})"
                    sheet.cell(row, column + 6).alignment = sheet.cell(row, column + 6).alignment.copy(horizontal="center")

                sheet.cell(row + 1, column=2 + x * (7 + 1)).value = "TOTAL"
                sheet.cell(row + 1, column=2 + x * (7 + 1)).font = sheet.cell(row + 1, column=2 + x * (7 + 1)).font.copy(bold=True)
                sheet.cell(row + 1, column+1).value = f"=SUM({get_column_letter(column+1)}{row-table_size+1}:{get_column_letter(column+1)}{row})"
                sheet.cell(row + 1, column+1).alignment = sheet.cell(row + 1, column+1).alignment.copy(horizontal="center")
                sheet.cell(row + 1, column+3).value = f"=SUM({get_column_letter(column+3)}{row-table_size+1}:{get_column_letter(column+3)}{row})"
                sheet.cell(row + 1, column+3).alignment = sheet.cell(row + 1, column+3).alignment.copy(horizontal="center")

    def fill_values(self, sheet, samples_number, group_id, number, table, track_id):
        areas = [float(x[-3]) for x in table]

        col = 3 + group_id * (7 + 1) + (0 if track_id % 2 == 0 else 2)
        row = 6 + ((number-1) % samples_number) * (len(table) + 4)

        for id, area in enumerate(areas):
            print(row + id, col)
            sheet.cell(row + id, col).value = area
            sheet.cell(row + id, col).alignment = sheet.cell(row + id, col).alignment.copy(horizontal="center")


    def build_excel(self, tables, configuration):
        groups = configuration["groups"]
        samples = configuration["samples"]

        if os.path.exists(self.output):
            wb = load_workbook(self.output)
        else:
            wb = Workbook()

        sheets = wb.sheetnames
        
        keys = sorted(list(tables.keys()))
        for id, k in enumerate(keys):
            substances = [x[-1] for x in tables[k]]
            print(k)

            organ, group, number = re.findall(r"Track \d+, ID: ([a-zA-Z ]*) ([a-zA-Z]+) *(\d+) \(.*", k)[0]
            
            if organ not in sheets:
                wb.create_sheet(organ)
                sheets.append(organ)
                self.build_sheet(wb[organ], groups, samples, substances)

            sheet = wb[organ]
            self.fill_values(sheet, samples, [x["acronym"] for x in groups].index(group), int(number), tables[k], id)

        if "Sheet" in sheets:
            wb.remove(wb["Sheet"])

        wb.save(self.output)


if __name__ == "__main__":
    reader = HptlcReader("C:/Users/Andre.LAPTOP01/Desktop/Projects/Helps/Koalinha/HPTLC_Creator/files/Rim_C5,6,7_16-11-23.pdf")
    tables = reader.extract_info()