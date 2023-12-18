from __future__ import annotations

import os

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import BORDER_MEDIUM
from openpyxl.styles.borders import Side

from openpyxl.utils.cell import get_column_letter

class ExcelWriter:
    def __init__(self, path: str) -> ExcelWriter:
        self.path = path
        if os.path.exists(path):
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()

    def save(self) -> None:
        """
        Saves the workbook.
        """
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        self.wb.save(self.path)

    def get_sheet(self, name: str) -> Worksheet:
        """
        Gets a sheet from the workbook.

        :param name: The name of the sheet.
        """
        if name in self.wb.sheetnames:
            return self.wb[name]
        else:
            return self.wb.create_sheet(name)
        
    def modify_cell(self, sheet: Worksheet, row: int, column: int, value: str, bold: bool = False, fg_color: str = "ffffff", alignment: str = "center") -> None:
        """
        Modifies a cell from a sheet.
        
        :param sheet: The sheet to modify.
        :param row: The row of the cell.
        :param column: The column of the cell.
        :param value: The value of the cell.
        :param bold: Whether the cell should be bold or not.
        :param color: The color of the cell.
        :param alignment: The alignment of the cell.
        """
        sheet.cell(row, column).value = value
        sheet.cell(row, column).font = sheet.cell(row, column).font.copy(bold=bold)
        sheet.cell(row, column).alignment = sheet.cell(row, column).alignment.copy(horizontal=alignment)
        sheet.cell(row, column).fill = PatternFill(start_color=fg_color,end_color=fg_color, fill_type='solid')
            

    def set_borders(self, sheet: Worksheet, row: int, column: int, border_color: str) -> None:
        """
        Set the borders of a cell from a sheet.

        :param sheet: The sheet to modify.
        :param row: The row of the cell.
        :param column: The column of the cell.
        :param border_color: The color of the border.
        """

        sheet.cell(row, column).border = Border(
            left=Side(border_style=BORDER_MEDIUM, color=border_color),
            right=Side(border_style=BORDER_MEDIUM, color=border_color),
            top=Side(border_style=BORDER_MEDIUM, color=border_color),
            bottom=Side(border_style=BORDER_MEDIUM, color=border_color)
        )


    def build_sheet_structure(self, name: str, data: dict, force=False) -> None:
        """
        Builds a sheet from a dictionary.

        :param name: The name of the sheet.
        :param data: The data necessarity to build the sheet (groups, samples, and substances).
        """
        if name in self.wb.sheetnames and not force: return

        sheet = self.get_sheet(name)
        groups, samples, substances = data["groups"], data["samples"], data["substances"]

        table_size = len(substances)
        for x in range(len(groups)):

            # Setup the group Header
            sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=2, end_row=2)
            self.modify_cell(sheet, 2, 2 + x * (7 + 1), groups[x]["name"].upper(), True, groups[x]["main_color"])

            # Setup the table header
            for y in range(samples):

                for col in range(2 + x * (7 + 1), 2 + x * (7 + 1) + 7):
                    for row in range(4 + y * (table_size + 4), 7 + table_size + y * (table_size + 4)):
                        self.set_borders(sheet, row, col, groups[x]["main_color"])

                sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=4 + y * (table_size + 4), end_row= 4 + y * (table_size + 4))
                self.modify_cell(sheet, 4 + y * (table_size + 4), 2 + x * (7 + 1), f"{groups[x]['acronym']}{x * samples + y + 1}", True, groups[x]["secondary_color"])
                self.set_borders(sheet, 4 + y * (table_size + 4), 2 + x * (7 + 1), groups[x]["main_color"])

                headers = ["Lípidos", "Área", "%", "Área", "%", "Média", "Desvio Padrão"]
                for i in range(len(headers)):
                    if i == 0:
                        self.modify_cell(sheet, 5 + y * (table_size + 4), 2 + x * (7 + 1) + i, headers[i], True, groups[x]["tertiary_color"], "left")
                    else:
                        self.modify_cell(sheet, 5 + y * (table_size + 4), 2 + x * (7 + 1) + i, headers[i], True, groups[x]["tertiary_color"])

                for id, s in enumerate(substances):
                    row = 6 + id + y * (table_size + 4)
                    column = 2 + x * (7 + 1)

                    self.modify_cell(sheet, row, column, s, alignment = "left")
                    self.modify_cell(sheet, row, column + 2, f"=(${get_column_letter(column+1)}{row}/${get_column_letter(column+1)}${row + table_size - id})*100", alignment = "center")
                    self.modify_cell(sheet, row, column + 4, f"=(${get_column_letter(column+3)}{row}/${get_column_letter(column+3)}{row + table_size - id})*100", alignment = "center")
                    self.modify_cell(sheet, row, column + 5, f"=AVERAGE(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})", alignment = "center")
                    self.modify_cell(sheet, row, column + 6, f"=_xlfn.STDEV.P(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})", alignment = "center")

                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1), value="TOTAL", bold=True)
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 1, value=f"=SUM({get_column_letter(column+1)}{row-table_size+1}:{get_column_letter(column+1)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 2, value=f"=SUM({get_column_letter(column+2)}{row-table_size+1}:{get_column_letter(column+2)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 3, value=f"=SUM({get_column_letter(column+3)}{row-table_size+1}:{get_column_letter(column+3)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 4, value=f"=SUM({get_column_letter(column+4)}{row-table_size+1}:{get_column_letter(column+4)}{row})", alignment = "center")


    def has_percentage_total(self, name: str) -> bool:
        """
        Checks if the sheet has a percentage total.

        :param name: The name of the sheet.
        :return: Whether the sheet has a percentage total or not.
        """
        sheet = self.get_sheet(name)

        c_value = sheet.cell(11, 4).value
        return c_value is not None and c_value.strip() != ""
    

    def fill_values(self, name: str, data: dict, project_config: dict) -> None:
        """
        Fill the HPTLC values in the sheet.

        :param name: The name of the sheet.
        :param data: The data necessarity to fill the sheet (samples, group_id, number, table, left_col).
        """
        if not self.has_percentage_total(name):
            self.build_sheet_structure(name, project_config, force = True)

        sheet = self.get_sheet(name)
        samples, group_id, number, table, left_col = data["samples"], data["group_id"], data["number"], data["table"], data["left_col"]

        areas = [float(x[-3]) for x in table]

        col = 3 + group_id * (7 + 1) + (0 if left_col else 2)
        row = 6 + ((number-1) % samples) * (len(table) + 4)

        for id, area in enumerate(areas):
            self.modify_cell(sheet, row + id, col, area)