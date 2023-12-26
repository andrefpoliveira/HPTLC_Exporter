from __future__ import annotations

import os

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import BORDER_MEDIUM
from openpyxl.styles.borders import Side

from abc import ABC, abstractmethod

class ExcelWriter(ABC):
    def __init__(self, path: str) -> ExcelWriter:
        self.path = path
        if os.path.exists(path):
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()

    def save(self) -> None:
        """
        Saves the workbook.
        """
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        self.wb.save(self.path)

    def get_sheet(self, name: str) -> Worksheet:
        """
        Gets a sheet from the workbook.

        :param name: The name of the sheet.
        """
        if name in self.wb.sheetnames:
            return self.wb[name]
        else:
            return self.wb.create_sheet(name)
        
    def modify_cell(self, sheet: Worksheet, row: int, column: int, value: str, bold: bool = False, fg_color: str = "ffffff", alignment: str = "center") -> None:
        """
        Modifies a cell from a sheet.
        
        :param sheet: The sheet to modify.
        :param row: The row of the cell.
        :param column: The column of the cell.
        :param value: The value of the cell.
        :param bold: Whether the cell should be bold or not.
        :param color: The color of the cell.
        :param alignment: The alignment of the cell.
        """
        sheet.cell(row, column).value = value
        sheet.cell(row, column).font = sheet.cell(row, column).font.copy(bold=bold)
        sheet.cell(row, column).alignment = sheet.cell(row, column).alignment.copy(horizontal=alignment)
        sheet.cell(row, column).fill = PatternFill(start_color=fg_color,end_color=fg_color, fill_type='solid')
            

    def set_borders(self, sheet: Worksheet, row: int, column: int, border_color: str) -> None:
        """
        Set the borders of a cell from a sheet.

        :param sheet: The sheet to modify.
        :param row: The row of the cell.
        :param column: The column of the cell.
        :param border_color: The color of the border.
        """

        sheet.cell(row, column).border = Border(
            left=Side(border_style=BORDER_MEDIUM, color=border_color),
            right=Side(border_style=BORDER_MEDIUM, color=border_color),
            top=Side(border_style=BORDER_MEDIUM, color=border_color),
            bottom=Side(border_style=BORDER_MEDIUM, color=border_color)
        )


    def has_percentage_total(self, name: str) -> bool:
        """
        Checks if the sheet has a percentage total.

        :param name: The name of the sheet.
        :return: Whether the sheet has a percentage total or not.
        """
        sheet = self.get_sheet(name)

        c_value = sheet.cell(11, 4).value
        return c_value is not None and c_value.strip() != ""
    
    @abstractmethod
    def build_sheet_structure(self, name: str, data: dict, force=False) -> None:
        """
        Builds a sheet from a dictionary.

        :param name: The name of the sheet.
        :param data: The data necessarity to build the sheet (groups, samples, and substances).
        """
        pass


    @abstractmethod
    def fill_values(self, name: str, data: dict, project_config: dict) -> None:
        """
        Fill the HPTLC values in the sheet.

        :param name: The name of the sheet.
        :param data: The data necessarity to fill the sheet (samples, group_id, number, table, left_col).
        """
        pass