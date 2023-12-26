from __future__ import annotations

from openpyxl.utils.cell import get_column_letter

from src.general.writer.excel_writer import ExcelWriter

class OtherExcelWriter(ExcelWriter):
    def __init__(self, path: str) -> OtherExcelWriter:
        super().__init__(path)

    def build_sheet_structure(self, data: dict, force=False) -> None:
        """
        Builds a sheet from a dictionary.

        :param name: The name of the sheet.
        :param data: The data necessarity to build the sheet (groups, samples, and substances).
        """
        if "HPTLC" in self.wb.sheetnames and not force: return

        sheet = self.get_sheet("HPTLC")
        groups, samples, substances = data["groups"], data["samples"], data["substances"]

        table_size = len(substances)
        samples_size = len(samples)
        for x in range(len(groups)):

            # Setup the group Header
            sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=2, end_row=2)
            self.modify_cell(sheet, 2, 2 + x * (7 + 1), groups[x]["name"].upper(), True, groups[x]["main_color"])

            # Setup the table header
            for y in range(samples_size):

                for col in range(2 + x * (7 + 1), 2 + x * (7 + 1) + 7):
                    for row in range(4 + y * (table_size + 4), 7 + table_size + y * (table_size + 4)):
                        self.set_borders(sheet, row, col, groups[x]["main_color"])

                sheet.merge_cells(start_column= 2 + x * (7 + 1), end_column= 2 + x * (7 + 1) + 6, start_row=4 + y * (table_size + 4), end_row= 4 + y * (table_size + 4))
                self.modify_cell(sheet, 4 + y * (table_size + 4), 2 + x * (7 + 1), f"{samples[y]}", True, groups[x]["secondary_color"])
                self.set_borders(sheet, 4 + y * (table_size + 4), 2 + x * (7 + 1), groups[x]["main_color"])

                headers = ["Lípidos", "Área", "%", "Área", "%", "Média", "Desvio Padrão"]
                for i in range(len(headers)):
                    if i == 0:
                        self.modify_cell(sheet, 5 + y * (table_size + 4), 2 + x * (7 + 1) + i, headers[i], True, groups[x]["tertiary_color"], "left")
                    else:
                        self.modify_cell(sheet, 5 + y * (table_size + 4), 2 + x * (7 + 1) + i, headers[i], True, groups[x]["tertiary_color"])

                for id, s in enumerate(substances):
                    row = 6 + id + y * (table_size + 4)
                    column = 2 + x * (7 + 1)

                    self.modify_cell(sheet, row, column, s.split("(")[0].strip(), alignment = "left")
                    self.modify_cell(sheet, row, column + 2, f"=(${get_column_letter(column+1)}{row}/${get_column_letter(column+1)}${row + table_size - id})*100", alignment = "center")
                    self.modify_cell(sheet, row, column + 4, f"=(${get_column_letter(column+3)}{row}/${get_column_letter(column+3)}{row + table_size - id})*100", alignment = "center")
                    self.modify_cell(sheet, row, column + 5, f"=AVERAGE(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})", alignment = "center")
                    self.modify_cell(sheet, row, column + 6, f"=_xlfn.STDEV.P(${get_column_letter(column+2)}{row},${get_column_letter(column+4)}{row})", alignment = "center")

                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1), value="TOTAL", bold=True)
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 1, value=f"=SUM({get_column_letter(column+1)}{row-table_size+1}:{get_column_letter(column+1)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 2, value=f"=SUM({get_column_letter(column+2)}{row-table_size+1}:{get_column_letter(column+2)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 3, value=f"=SUM({get_column_letter(column+3)}{row-table_size+1}:{get_column_letter(column+3)}{row})", alignment = "center")
                self.modify_cell(sheet, row + 1, column=2 + x * (7 + 1) + 4, value=f"=SUM({get_column_letter(column+4)}{row-table_size+1}:{get_column_letter(column+4)}{row})", alignment = "center")
    
    def fill_values(self, data: dict, project_config: dict) -> None:
        """
        Fill the HPTLC values in the sheet.

        :param name: The name of the sheet.
        :param data: The data necessarity to fill the sheet (samples, group_id, number, table, left_col).
        """
        if not self.has_percentage_total("HPTLC"):
            self.build_sheet_structure("HPTLC", project_config, force = True)

        sheet = self.get_sheet("HPTLC")
        samples, group_id, sample, table, left_col = data["samples"], data["group_id"], data["sample"], data["table"], data["left_col"]

        areas = [float(x[-3]) for x in table]

        col = 3 + group_id * (7 + 1) + (0 if left_col else 2)
        row = 6 + samples.index(sample.strip()) * (len(table) + 4)

        for id, area in enumerate(areas):
            self.modify_cell(sheet, row + id, col, area)